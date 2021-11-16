from openpyxl import load_workbook
import openpyxl
print(openpyxl.__version__)
wb = load_workbook('demo.xlsx')
# print(wb.sheetnames)
ws = wb.worksheets[0]

img = openpyxl.drawing.image.Image('01-001_1.jpg')
img.anchor = 'B5'
ws.add_image(img)

print(ws.max_row)

wb.save('demo.xlsx')

